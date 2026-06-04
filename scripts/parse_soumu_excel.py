"""
総務省現況調査Excelパーサー
使い方: python scripts/parse_soumu_excel.py
出力: data/jichitai_detail.csv（2024年度）
      data/jichitai_detail_multi.csv（2019〜2024年度 費目別経費）
      data/jichitai_timeseries.csv（2008〜2024年度 受入額時系列）
"""
import pandas as pd
from pathlib import Path

DATA_DIR = Path(__file__).parent.parent / "data"
RAW_DIR = DATA_DIR / "raw"

# 年度別ファイル設定（ファイル名: (年度, シート名override)）
DETAIL_FILES = {
    "soumu_R1_2019detail.xlsx": (2019, None),
    "soumu_R2_2020detail.xlsx": (2020, None),
    "soumu_R3_2021detail.xlsx": (2021, None),
    "soumu_R4_2022detail.xlsx": (2022, "Sheet1"),
    "soumu_R5_2023detail.xlsx": (2023, None),
    "001022818_2024detail.xlsx": (2024, None),
}

# 列の定義（全年度共通・0始まり）
COL_MAP = {
    0:  "団体コード",
    1:  "都道府県",
    2:  "市区町村",
    3:  "受入件数",
    4:  "受入額_円",
    11: "返礼品調達費_円",
    12: "送付費_円",
    13: "広報費_円",
    14: "決済費_円",
    15: "事務費_円",
    16: "その他費_円",
    17: "経費合計_円",
    18: "返礼品率",
    19: "送付費率",
    20: "広報費率",
    21: "決済費率",
    22: "事務費率",
    24: "経費率合計",
}
# ポータル費は2024年度のみ列25に存在
COL_MAP_2024 = {**COL_MAP, 25: "ポータル費_円"}


def find_data_start(xlsx_path: Path, sheet_name) -> int:
    """団体コード（6桁数字）が始まる行番号を自動検出"""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        if row[0] and str(row[0]).strip().isdigit() and len(str(row[0]).strip()) == 6:
            wb.close()
            return i
    wb.close()
    return 15  # デフォルト


def get_sheet_name(xlsx_path: Path, override) -> str:
    if override:
        return override
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    for s in names:
        if "集計" in s:
            return s
    return names[0]


def parse_detail(xlsx_path: Path, year: int, sheet_override=None) -> pd.DataFrame:
    """費目別経費明細Excelを読み込む（全年度対応）"""
    sheet = get_sheet_name(xlsx_path, sheet_override)
    data_start = find_data_start(xlsx_path, sheet)
    skiprows = data_start - 1

    col_map = COL_MAP_2024 if year == 2024 else COL_MAP
    df = pd.read_excel(xlsx_path, sheet_name=sheet, header=None,
                       skiprows=skiprows, dtype=str)

    # 使用する列のみ抽出（存在する列のみ）
    valid_cols = {k: v for k, v in col_map.items() if k < len(df.columns)}
    df = df[list(valid_cols.keys())].rename(columns=valid_cols)

    # 2024年度以外はポータル費列なし
    if "ポータル費_円" not in df.columns:
        df["ポータル費_円"] = None

    # 団体コードが6桁数字の行だけ残す
    df = df[df["団体コード"].str.match(r"^\d{6}$", na=False)].copy()

    # 数値変換
    num_cols = [c for c in df.columns if c not in ("都道府県", "市区町村", "団体コード")]
    for c in num_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    df["年度"] = year
    df["受入額_億円"]  = (df["受入額_円"]   / 1e8).round(2)
    df["経費合計_億円"] = (df["経費合計_円"] / 1e8).round(2)
    df["ポータル費_億円"] = (df["ポータル費_円"] / 1e8).round(2)

    return df


def parse_timeseries(xlsx_path: Path) -> pd.DataFrame:
    """001022819形式（各団体一覧・時系列）を読み込む"""
    raw = pd.read_excel(xlsx_path, sheet_name="各団体一覧", header=None, dtype=str)

    year_row = raw.iloc[1].tolist()
    years = [str(v).strip() if pd.notna(v) and str(v) not in ["None", ""] else None
             for v in year_row[2:]]
    filled_years = []
    last = None
    for y in years:
        if y:
            last = y
        filled_years.append(last)

    records = []
    for _, row in raw.iloc[4:].iterrows():
        row = row.tolist()
        pref = str(row[0]).strip() if pd.notna(row[0]) else None
        muni = str(row[1]).strip() if pd.notna(row[1]) and str(row[1]) not in ["None", "nan"] else None
        if not pref or pref in ["None", "nan"]:
            continue
        data_vals = row[2:]
        for i in range(0, len(filled_years) - 1, 2):
            if i >= len(data_vals):
                break
            year_label = filled_years[i]
            if not year_label:
                continue
            try:
                kin = float(data_vals[i]) if data_vals[i] not in [None, "None", "nan", ""] else None
                ken = float(data_vals[i+1]) if i+1 < len(data_vals) and data_vals[i+1] not in [None, "None", "nan", ""] else None
            except (ValueError, TypeError):
                kin, ken = None, None
            records.append({"都道府県": pref, "市区町村": muni or pref,
                            "年度ラベル": year_label, "受入額_千円": kin, "受入件数": ken})

    df = pd.DataFrame(records)
    mapping = {
        "平成20年度": 2008, "平成21年度": 2009, "平成22年度": 2010, "平成23年度": 2011,
        "平成24年度": 2012, "平成25年度": 2013, "平成26年度": 2014, "平成27年度": 2015,
        "平成28年度": 2016, "平成29年度": 2017, "平成30年度": 2018, "令和元年度": 2019,
        "令和２年度": 2020, "令和３年度": 2021, "令和４年度": 2022, "令和５年度": 2023,
        "令和６年度": 2024,
    }
    df["年度"] = df["年度ラベル"].map(mapping)
    df["受入額_億円"] = (df["受入額_千円"] / 1e5).round(2)
    df = df.dropna(subset=["年度", "受入額_千円"])
    return df[["都道府県", "市区町村", "年度", "受入額_億円", "受入件数"]].sort_values(
        ["都道府県", "市区町村", "年度"])


if __name__ == "__main__":
    dfs = []

    # 全年度の費目別明細をパース
    for fname, (year, sheet_override) in DETAIL_FILES.items():
        path = RAW_DIR / fname
        if not path.exists():
            print(f"スキップ（ファイルなし）: {fname}")
            continue
        print(f"パース中: {fname} ({year}年度)...")
        df = parse_detail(path, year=year, sheet_override=sheet_override)
        dfs.append(df)
        print(f"  → {len(df)}自治体")

    if dfs:
        # 最新年度（2024）を単体でも保存
        df_latest = [d for d in dfs if d["年度"].iloc[0] == 2024]
        if df_latest:
            out = DATA_DIR / "jichitai_detail.csv"
            df_latest[0].to_csv(out, index=False, encoding="utf-8-sig")
            print(f"\n保存: {out} ({len(df_latest[0])}行)")

        # 全年度を結合して保存
        df_multi = pd.concat(dfs, ignore_index=True).sort_values(["年度", "都道府県", "市区町村"])
        out_multi = DATA_DIR / "jichitai_detail_multi.csv"
        df_multi.to_csv(out_multi, index=False, encoding="utf-8-sig")
        print(f"保存: {out_multi} ({len(df_multi)}行 / {df_multi['年度'].nunique()}年度分)")

    # 時系列
    ts_path = RAW_DIR / "001022819_timeseries.xlsx"
    if ts_path.exists():
        print("\nパース中: 時系列...")
        df_ts = parse_timeseries(ts_path)
        out = DATA_DIR / "jichitai_timeseries.csv"
        df_ts.to_csv(out, index=False, encoding="utf-8-sig")
        print(f"保存: {out} ({len(df_ts)}行)")

    print("\n完了!")
